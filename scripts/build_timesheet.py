#!/usr/bin/env python3
"""
Builds output/timesheet.xlsx from data/punch_records.csv.

Run this after new rows are appended to the CSV:
    python3 scripts/build_timesheet.py
"""
import csv
import datetime as dt
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
CSV_PATH = ROOT / "data" / "punch_records.csv"
OUT_PATH = ROOT / "output" / "timesheet.xlsx"

HEADERS = ["Employee ID", "Date", "Time Out", "Time In", "Total Hours", "Notes"]
FONT_NAME = "Arial"
BLANK = "-"
FRIDAY_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def parse_date(value):
    value = (value or "").strip()
    if not value:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return dt.datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return value  # leave as raw text if unparseable, so nothing silently disappears


def parse_time(value):
    value = (value or "").strip()
    if not value:
        return None
    for fmt in ("%H:%M", "%H:%M:%S", "%I:%M %p", "%I:%M:%S %p"):
        try:
            return dt.datetime.strptime(value, fmt).time()
        except ValueError:
            continue
    return value


def calc_total_hours(time_in_val, time_out_val):
    if not isinstance(time_in_val, dt.time) or not isinstance(time_out_val, dt.time):
        return None
    if time_in_val == time_out_val:
        return None  # identical in/out is a flagged read error, not a real 0/24h shift
    start = dt.datetime.combine(dt.date.min, time_in_val)
    end = dt.datetime.combine(dt.date.min, time_out_val)
    if end <= start:
        end += dt.timedelta(days=1)  # shift crossed midnight (e.g. night security)
    return round((end - start).total_seconds() / 3600, 2)


def sort_key(rec):
    try:
        emp_id = int(rec.get("employee_id", "") or 0)
    except ValueError:
        emp_id = rec.get("employee_id", "")
    date_val = parse_date(rec.get("date", ""))
    date_key = date_val if isinstance(date_val, dt.date) else dt.date.min
    return (emp_id, date_key)


def fill_missing_days(rows):
    all_dates = [parse_date(rec.get("date", "")) for rec in rows]
    all_dates = [d for d in all_dates if isinstance(d, dt.date)]
    if not all_dates:
        return rows

    full_range = [
        dt.date.fromordinal(o)
        for o in range(min(all_dates).toordinal(), max(all_dates).toordinal() + 1)
    ]

    by_employee = {}
    for rec in rows:
        by_employee.setdefault(rec.get("employee_id", ""), set()).add(
            parse_date(rec.get("date", ""))
        )

    filled = list(rows)
    for employee_id, existing_dates in by_employee.items():
        for day in full_range:
            if day not in existing_dates:
                filled.append(
                    {
                        "employee_id": employee_id,
                        "date": day.isoformat(),
                        "time_in": "",
                        "time_out": "",
                        "notes": "No entry",
                    }
                )
    return filled


def main():
    with open(CSV_PATH, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    rows = fill_missing_days(rows)
    rows.sort(key=sort_key)

    wb = Workbook()
    ws = wb.active
    ws.title = "Punch Records"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF")
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    r = 2
    prev_employee_id = None
    for rec in rows:
        employee_id = rec.get("employee_id", "")
        if prev_employee_id is not None and employee_id != prev_employee_id:
            r += 1  # blank row separating employees
        prev_employee_id = employee_id

        date_val = parse_date(rec.get("date", "")) or BLANK
        time_out_raw = parse_time(rec.get("time_out", ""))
        time_in_raw = parse_time(rec.get("time_in", ""))
        total_hours = calc_total_hours(time_in_raw, time_out_raw)
        time_out_val = time_out_raw or BLANK
        time_in_val = time_in_raw or BLANK
        notes_val = rec.get("notes", "").strip() or BLANK
        values = [
            employee_id or BLANK,
            date_val,
            time_out_val,
            time_in_val,
            total_hours if total_hours is not None else BLANK,
            notes_val,
        ]
        is_friday = isinstance(date_val, dt.date) and date_val.weekday() == 4
        for c, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name=FONT_NAME)
            if c == 2 and isinstance(val, dt.date):
                cell.number_format = "yyyy-mm-dd"
            if c in (3, 4) and isinstance(val, dt.time):
                cell.number_format = "HH:MM"
            if c == 5 and isinstance(val, (int, float)):
                cell.number_format = "0.00"
            if is_friday:
                cell.fill = FRIDAY_FILL
        r += 1

    widths = [14, 12, 10, 10, 12, 40]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {len(rows)} rows to {OUT_PATH}")


if __name__ == "__main__":
    main()
